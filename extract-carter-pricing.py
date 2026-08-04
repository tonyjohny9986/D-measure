import json
from pathlib import Path

import openpyxl


WORKBOOK_PATH = Path("/Users/tonyz/Desktop/Desktop - Mac/Decovibes/oder sent/Copy of Carter Blinds Digital Price List V2.xlsx")
OUTPUT_PATH = Path("/Users/tonyz/Downloads/DOWNLOAD-FIXED/data/carter-pricing.json")


def text(value):
    return str(value).strip() if value is not None else ""


def number(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def matrix_key_from_title(title):
    label = text(title)
    if label.startswith("Category "):
        label = label[len("Category ") :]
    return label


def parse_matrix_sections(ws):
    sections = {}
    row = 1
    while row <= ws.max_row:
        title = ws.cell(row, 1).value
        if isinstance(title, str) and text(title).startswith("Category "):
            key = matrix_key_from_title(title)
            width_row = row + 1
            widths = []
            col = 2
            while col <= ws.max_column:
                value = number(ws.cell(width_row, col).value)
                if value is None:
                    break
                widths.append(int(round(value)))
                col += 1

            heights = []
            prices = []
            data_row = row + 2
            while data_row <= ws.max_row:
                height = number(ws.cell(data_row, 1).value)
                if height is None:
                    break
                row_prices = []
                valid = True
                for offset in range(len(widths)):
                    price = number(ws.cell(data_row, 2 + offset).value)
                    if price is None:
                        valid = False
                        break
                    row_prices.append(round(price, 6))
                if not valid:
                    break
                heights.append(int(round(height)))
                prices.append(row_prices)
                data_row += 1

            sections[key] = {
                "widthsMm": widths,
                "dropsMm": heights,
                "prices": prices,
            }
            row = data_row
            continue
        row += 1
    return sections


def parse_single_matrix(ws, key_name):
    width_row = None
    for row in range(1, ws.max_row + 1):
        if number(ws.cell(row, 2).value) is not None:
            width_row = row
            break
    if width_row is None:
        return {}

    widths = []
    col = 2
    while col <= ws.max_column:
        value = number(ws.cell(width_row, col).value)
        if value is None:
            break
        widths.append(int(round(value)))
        col += 1

    drops = []
    prices = []
    row = width_row + 1
    while row <= ws.max_row:
        drop = number(ws.cell(row, 1).value)
        if drop is None:
            break
        row_prices = []
        valid = True
        for offset in range(len(widths)):
            price = number(ws.cell(row, 2 + offset).value)
            if price is None:
                valid = False
                break
            row_prices.append(round(price, 6))
        if not valid:
            break
        drops.append(int(round(drop)))
        prices.append(row_prices)
        row += 1

    return {
        key_name: {
            "widthsMm": widths,
            "dropsMm": drops,
            "prices": prices,
        }
    }


def parse_offset_matrix(ws, width_row, width_start_col, drop_start_row, drop_col):
    widths = []
    col = width_start_col
    while col <= ws.max_column:
        value = number(ws.cell(width_row, col).value)
        if value is None:
            break
        widths.append(int(round(value)))
        col += 1

    drops = []
    prices = []
    row = drop_start_row
    while row <= ws.max_row:
        drop = number(ws.cell(row, drop_col).value)
        if drop is None:
            break
        row_prices = []
        valid = True
        for offset in range(len(widths)):
            price = number(ws.cell(row, width_start_col + offset).value)
            if price is None:
                valid = False
                break
            row_prices.append(round(price, 6))
        if not valid:
            break
        drops.append(int(round(drop)))
        prices.append(row_prices)
        row += 1

    return {
        "widthsMm": widths,
        "dropsMm": drops,
        "prices": prices,
    }


def add_fabric(fabrics, seen, name, supplier, category_key):
    fabric_name = text(name)
    if not fabric_name:
        return
    key = fabric_name.lower()
    if key in seen:
        return
    seen.add(key)
    fabrics.append(
        {
            "name": fabric_name,
            "supplier": text(supplier) or None,
            "categoryKey": category_key,
        }
    )


def parse_grid_categories(ws, start_rows, category_key_parser):
    fabrics = []
    seen = set()
    for start_row in start_rows:
        categories = []
        col = 1
        while col <= ws.max_column:
            label = text(ws.cell(start_row, col).value)
            if label:
                categories.append((col, category_key_parser(label)))
            col += 2
        row = start_row + 2
        while row <= ws.max_row:
            has_values = False
            for col, category_key in categories:
                range_name = text(ws.cell(row, col).value)
                supplier = text(ws.cell(row, col + 1).value)
                if range_name:
                    has_values = True
                    add_fabric(fabrics, seen, range_name, supplier, category_key)
            if not has_values:
                break
            row += 1
    return fabrics


def parse_price_text(value):
    raw = text(value)
    if not raw:
        return None
    cleaned = raw.replace("$", "").strip()
    if "%" in cleaned:
        pct = number(cleaned.replace("%", "").replace("more", "").strip())
        if pct is not None:
            return {"type": "percent", "percent": pct, "raw": raw}
    amount = number(cleaned.split()[0])
    if amount is not None:
        if "per bracket" in raw.lower() or "per blade" in raw.lower():
            return {"type": "manual", "raw": raw}
        return {"type": "fixed", "amount": round(amount, 6), "raw": raw}
    return {"type": "manual", "raw": raw}


def parse_simple_extras(ws):
    fixed = []
    percent = []
    manual = []
    row = 1
    while row <= ws.max_row:
        name = text(ws.cell(row, 1).value)
        value = ws.cell(row, 2).value
        if not name:
            row += 1
            continue
        if row == 1:
            row += 1
            continue
        if value is None:
            row += 1
            continue
        parsed = parse_price_text(value) if isinstance(value, str) else {"type": "fixed", "amount": round(number(value), 6), "raw": value}
        if not parsed:
            row += 1
            continue
        item = {"name": name}
        if parsed["type"] == "fixed":
            item["amount"] = parsed["amount"]
            fixed.append(item)
        elif parsed["type"] == "percent":
            item["percent"] = parsed["percent"]
            percent.append(item)
        else:
            item["pricingText"] = parsed["raw"]
            manual.append(item)
        row += 1
    return {"fixed": fixed, "percent": percent, "manual": manual}


def parse_motorisation(ws):
    groups = []
    current = None
    for row in range(1, ws.max_row + 1):
        name = text(ws.cell(row, 1).value)
        price = number(ws.cell(row, 2).value)
        if not name:
            continue
        if row == 1:
            continue
        if price is None:
            current = {"group": name, "options": []}
            groups.append(current)
            continue
        if current is None:
            current = {"group": "General", "options": []}
            groups.append(current)
        current["options"].append({"name": name, "amount": round(price, 6)})
    return groups


wb_values = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
wb_formulas = openpyxl.load_workbook(WORKBOOK_PATH, data_only=False)

roller_matrices = parse_matrix_sections(wb_values["Roller Blind Price List"])
vertical_matrices = parse_matrix_sections(wb_values["Vertical Blind Price List"])
verishade_matrices = parse_matrix_sections(wb_values["Verishade Price List"])
soft_wave_matrices = parse_single_matrix(wb_values["Soft Wave Price List"], "Soft Wave")

roller_sheet = wb_formulas["Roller&Panel Fabric Categories"]
roller_fabrics = []
roller_seen = set()
add_fabric(roller_fabrics, roller_seen, "Quest", None, "Quest")
add_fabric(roller_fabrics, roller_seen, "Vibe", None, "Vibe and Macie")
add_fabric(roller_fabrics, roller_seen, "Macie", None, "Vibe and Macie")
add_fabric(roller_fabrics, roller_seen, "Adventus 5%", None, "Adventus 5%")
roller_fabrics.extend(parse_grid_categories(roller_sheet, [8, 30], lambda label: text(label)))

vertical_sheet = wb_formulas["Vertical Fabric Categories"]
vertical_fabrics = []
vertical_seen = set()
add_fabric(vertical_fabrics, vertical_seen, "Quest", None, "Quest")
add_fabric(vertical_fabrics, vertical_seen, "Vibe", None, "Vibe")
vertical_fabrics.extend(parse_grid_categories(vertical_sheet, [6, 30], lambda label: text(label)))

verishade_sheet = wb_formulas["Soft Wave&Verishade Categories"]
verishade_fabrics = []
verishade_seen = set()
add_fabric(verishade_fabrics, verishade_seen, "Classic", None, "Classic")
verishade_fabrics.extend(parse_grid_categories(verishade_sheet, [6], lambda label: matrix_key_from_title(label)))

roller_extras = parse_simple_extras(wb_values["Roller Blind Extras"])
roller_extras["matrixFixed"] = [
    {
        "name": "Cassette",
        "matrix": parse_offset_matrix(wb_values["Roller Blind Extras"], 4, 6, 5, 5),
    }
]
roller_motorisation = parse_motorisation(wb_values["Roller Blind Motorisation"])
vertical_extras = parse_simple_extras(wb_values["Vertical Blind Extras"])

catalog = {
    "sourceWorkbook": str(WORKBOOK_PATH),
    "products": {
        "rollerBlind": {
            "displayName": "Roller Blind",
            "defaultFabric": "Vibe",
            "categories": roller_matrices,
            "fabrics": roller_fabrics,
            "extras": roller_extras,
            "motorisation": roller_motorisation,
        },
        "verticalBlind": {
            "displayName": "Vertical Blind",
            "defaultFabric": "Vibe",
            "categories": vertical_matrices,
            "fabrics": vertical_fabrics,
            "extras": vertical_extras,
        },
        "verishade": {
            "displayName": "Verishade",
            "defaultFabric": "Classic",
            "categories": verishade_matrices,
            "fabrics": verishade_fabrics,
        },
        "softWave": {
            "displayName": "Soft Wave",
            "defaultFabric": "Soft Wave",
            "categories": soft_wave_matrices,
            "fabrics": [
                {
                    "name": "Soft Wave",
                    "supplier": None,
                    "categoryKey": "Soft Wave",
                }
            ],
        },
    },
}

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH.write_text(json.dumps(catalog, indent=2), encoding="utf-8")
print(OUTPUT_PATH)
